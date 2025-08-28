import openpyxl
import os
import glob

def WFS_upload():
    folder = r"C:\ACT\公用核心\Walmart\WFS"

    # Search all .xlsx files under the folder (including subfolders)
    files = glob.glob(os.path.join(folder, "**", "*.xlsx"), recursive=True)

    if not files:
        raise FileNotFoundError(f"No .xlsx file found in {folder}")


    # Get the most recently modified file
    src_file = max(files, key=os.path.getmtime)
    wb = openpyxl.load_workbook(src_file)

    # Select the required sheet
    sheet_name = "正式Inbound"
    ws = wb[sheet_name]

    # Create a new workbook and copy data
    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active
    new_ws.title = "Create inbound order template"

    for row in ws.iter_rows(values_only=True):
        new_ws.append(row)

    # Get user's Downloads folder
    downloads_path = os.path.join(os.path.expanduser("~"), "Downloads")

    # Build export file name based on original file name
    base_name = os.path.splitext(os.path.basename(src_file))[0]
    output_file = os.path.join(downloads_path, f"{base_name}_export.xlsx")

    # Save to Downloads
    new_wb.save(output_file)
    print("✅ File exported successfully to Downloads")
