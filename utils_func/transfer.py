import os
import sys
import openpyxl
import tkinter as tk
from tkinter import simpledialog, messagebox
from datetime import datetime

# Add current directory to sys.path to import local modules
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from utils_func import mapping
from utils_func.excel_utils import save_as_xls  

def stock_transfer():
    root = tk.Tk()
    root.withdraw()

    # User input
    user_input = simpledialog.askstring("Data Input", "Please paste the data (three columns):")
    if not user_input:
        messagebox.showerror("Error", "No input detected.")
        return

    data = []
    for line in user_input.strip().splitlines():
        parts = line.strip().split()
        if len(parts) != 3:
            continue  # Skip invalid rows

        code = parts[0].strip().upper()  # Map only the first column
        sku = mapping.code_map.get(code)
        if sku:
            data.append([sku, parts[1], parts[2]])

    if not data:
        messagebox.showerror("Error", "No valid mapped data found.")
        return

    # Load template
    template_path = r"C:\Template\调仓.xlsx"
    if not os.path.exists(template_path):
        messagebox.showerror("Error", f"Template file not found: {template_path}")
        return

    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Write data to Excel
    start_row = 2
    for i, row in enumerate(data, start=start_row):
        ws.cell(row=i, column=1).value = row[0]  # Mapped SKU
        ws.cell(row=i, column=2).value = row[1]  # Original second column
        ws.cell(row=i, column=5).value = row[2]  # Original third column

    # Save XLSX
    today_str = datetime.today().strftime("%Y%m%d")
    output_dir = r"C:\Users\monica\Downloads"
    os.makedirs(output_dir, exist_ok=True)
    file_name_xlsx = f"StockTransfer_{today_str}.xlsx"
    output_path_xlsx = os.path.join(output_dir, file_name_xlsx)
    wb.save(output_path_xlsx)

    # Convert to XLS
    file_name_xls = f"StockTransfer_{today_str}.xls"
    output_path_xls = os.path.join(output_dir, file_name_xls)
    save_as_xls(output_path_xlsx, output_path_xls)

    # Delete XLSX file
    try:
        os.remove(output_path_xlsx)
    except Exception as e:
        messagebox.showwarning("Warning", f"Failed to delete XLSX file: {e}")

    messagebox.showinfo("Success", f"✅ Stock transfer file has been successfully created:\n{output_path_xls}")

'''
if __name__ == "__main__":
    stock_transfer()
'''
