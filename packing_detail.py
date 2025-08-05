import openpyxl
import tkinter as tk
from tkinter import simpledialog, messagebox
from datetime import datetime
from excel_utils import save_as_xls  
import os

def packing_detail():
    # Popup input for multiline data
    root = tk.Tk()
    root.withdraw()
    user_input = simpledialog.askstring("Data Input", "Please paste the data:\n(Use Tab to separate columns)")

    if not user_input:
        messagebox.showerror("Error", "No input detected.")
        return

    # Parse pasted data
    data = []
    for line in user_input.strip().splitlines():
        row = line.strip().split("\t")
        # 确保有足够的列（至少14列）
        if len(row) >= 14:
            a_col = row[13]  # 第14列 → 模板第1列
            b_col = row[1]   # 第2列 → 模板第2列
            c_col = row[3]   # 第4列 → 模板第3列
            data.append([a_col, b_col, c_col])

    if not data:
        messagebox.showerror("Error", "Parsed data is empty or does not contain enough columns.")
        return

    # Load template
    template_path = r"C:\Template\Packing_details.xlsx"
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Write data to worksheet
    start_row = 2
    for i, row in enumerate(data, start=start_row):
        ws.cell(row=i, column=1).value = row[0]  # A列：第14列
        ws.cell(row=i, column=2).value = row[1]  # B列：第2列
        ws.cell(row=i, column=3).value = row[2]  # C列：第4列

    # Save as xlsx
    today_str = datetime.today().strftime("%Y%m%d")
    file_name_xlsx = f"packing_{today_str}.xlsx"
    output_dir = r"C:\Users\monica\Downloads"
    output_path_xlsx = os.path.join(output_dir, file_name_xlsx)
    wb.save(output_path_xlsx)

    # Convert xlsx to xls
    file_name_xls = f"packing_{today_str}.xls"
    output_path_xls = os.path.join(output_dir, file_name_xls)
    save_as_xls(output_path_xlsx, output_path_xls)

    # Delete xlsx
    try:
        os.remove(output_path_xlsx)
    except Exception as e:
        messagebox.showwarning("Warning", f"Failed to delete xlsx file: {e}")

    # Done
    messagebox.showinfo("Success", f"✅ Data has been successfully saved as:\n{output_path_xls}")


if __name__ == "__main__":
    packing_detail()
