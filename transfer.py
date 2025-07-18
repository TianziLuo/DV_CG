import openpyxl
import tkinter as tk
from tkinter import simpledialog, messagebox
from datetime import datetime
from excel_utils import save_as_xls  
import os

def stock_transfer():

    # Popup input for multiline data
    root = tk.Tk()
    root.withdraw()
    user_input = simpledialog.askstring("Data Input", "Please paste the data:")

    if not user_input:
        messagebox.showerror("Error", "No input detected.")
        exit()

    data = []
    for line in user_input.strip().splitlines():
        parts = line.strip().split()
        if len(parts) == 3:
            try:
                parts[2] = int(parts[2])
                data.append(parts)
            except ValueError:
                print(f"⚠️ The third column is not a number: {line}")
        else:
            print(f"⚠️ Skipping invalid formatted line: {line}")

    if not data:
        messagebox.showerror("Error", "No valid data found.")
        exit()

    template_path = r"C:\Template\调仓.xlsx"
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    start_row = 2
    for i, row in enumerate(data, start=start_row):
        ws.cell(row=i, column=1).value = row[0]
        ws.cell(row=i, column=5).value = row[1]
        ws.cell(row=i, column=2).value = row[2]

    today_str = datetime.today().strftime("%Y%m%d")
    file_name_xlsx = f"调仓_{today_str}.xlsx"
    output_dir = r"C:\Users\monica\Downloads"
    output_path_xlsx = os.path.join(output_dir, file_name_xlsx)

    wb.save(output_path_xlsx)

    file_name_xls = f"调仓_{today_str}.xls"
    output_path_xls = os.path.join(output_dir, file_name_xls)

    # Convert xlsx to xls
    save_as_xls(output_path_xlsx, output_path_xls)

    # Delete the xlsx file
    try:
        os.remove(output_path_xlsx)
    except Exception as e:
        messagebox.showwarning("Warning", f"Failed to delete xlsx file: {e}")

    messagebox.showinfo("Success", f"✅ Data has been successfully saved as:\n{output_path_xls}")
