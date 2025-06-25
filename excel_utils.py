# excel_utils.py
from openpyxl import load_workbook
import xlwt
import time
from pathlib import Path
from typing import Tuple, List

# saved in 30s
def verify_file_fresh(path: Path, max_age_seconds: int = 30) -> None:
    if not path.exists():
        raise FileNotFoundError(f"File not found: {path}")
    age = time.time() - path.stat().st_mtime
    if age > max_age_seconds:
        raise TimeoutError(f"{path.name} is {int(age)} s old (> {max_age_seconds}s).")


def load_worksheet(workbook_path: Path, sheet_name: str):
    wb = load_workbook(workbook_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise KeyError(f"Sheet '{sheet_name}' not found in {workbook_path}")
    return wb, wb[sheet_name]

def find_last_data_row(ws, col_range: Tuple[int, int], start_row: int = 1) -> int:
    start_col, end_col = col_range
    last_row = ws.max_row
    while last_row >= start_row and all(
        ws.cell(row=last_row, column=col).value is None
        for col in range(start_col, end_col + 1)
    ):
        last_row -= 1
    return last_row

# copy range, src_ws -> tgt_ws
def copy_range_to_ws(
    src_ws,
    tgt_ws,
    row_range: Tuple[int, int],
    col_map: List[Tuple[int, int]],
):
    for src_row in range(*row_range):
        tgt_row = src_row - row_range[0] + 2  # Paste from row 2
        for src_col, tgt_col in col_map:
            val = src_ws.cell(row=src_row, column=src_col).value
            tgt_ws.cell(row=tgt_row, column=tgt_col).value = val

# xlsx -> xls
def save_as_xls(xlsx_path: Path, xls_path: Path) -> None:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    wb_xls = xlwt.Workbook()
    ws_xls = wb_xls.add_sheet("Sheet1")
    for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
        for c_idx, val in enumerate(row):
            ws_xls.write(r_idx, c_idx, "" if val is None else val)
    wb_xls.save(str(xls_path))
